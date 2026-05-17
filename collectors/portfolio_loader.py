"""포트폴리오 정리.xlsx 자동 파서.

한 셀에 여러 종목이 줄바꿈으로 들어있는 형태를 정규식으로 파싱.
yfinance 추적 가능 종목은 ticker_map.csv로 티커 매핑.
펀드·연금·현금성은 별도 카테고리로 분류.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

from config import ROOT, get_logger

log = get_logger(__name__)

EXCEL_PATH = ROOT / "포트폴리오 정리.xlsx"
TICKER_MAP_PATH = ROOT / "ticker_map.csv"
UNMATCHED_LOG = ROOT / "logs" / "unmatched.csv"


# 카테고리 분류
TRADEABLE_CATEGORIES = {"주식", "ISA", "IRP", "퇴직연금(DC)"}
BOND_CATEGORIES = {"채권"}
FUND_CATEGORIES = {"변액연금보험", "VUL", "변액종신보험"}
CASH_CATEGORIES = {"RP", "CMA"}
PENSION_CATEGORIES = {"개인연금", "국민연금"}


@dataclass
class Holding:
    no: int
    category: str
    provider: str
    name: str
    raw: str  # 원본 줄
    quantity: float | None = None
    value_krw: float | None = None
    return_pct: float | None = None
    ticker: str | None = None
    asset_type: str = "unknown"  # stock, etf, bond, fund, cash, pension, insurance, unknown
    currency: str = "KRW"
    notes: list[str] = field(default_factory=list)


# ──────────────────────────────────────────────────────────────────
# 정규식
# ──────────────────────────────────────────────────────────────────
RE_QUANTITY = re.compile(r"(\d{1,3}(?:,\d{3})+|\d+)\s*(주|좌)")
RE_VALUE_OK = re.compile(r"(\d+(?:\.\d+)?)\s*억(?:원)?")
RE_VALUE_MAN = re.compile(r"(\d+(?:[\.,]\d+)?)\s*만원")
RE_VALUE_WON = re.compile(r"(\d{4,}(?:,\d{3})*)\s*원")
RE_RETURN = re.compile(r"\(?([+\-]?\d+(?:\.\d+)?)\s*%\)?")
RE_BOND_GUKGO = re.compile(r"^(국고\S*)\s+(\d+\.\d+)\s*억\s+([+\-]?\d+(?:\.\d+)?)\s*억", re.M)
RE_BOND_US = re.compile(r"^(미국채\S*[^0-9]+\S*\s+\d+/\d+/\d+)\s+(\d+\.\d+)\s*억\s+([+\-]?\d+(?:\.\d+)?)\s*억", re.M)


def _parse_value(text: str) -> float | None:
    """텍스트에서 평가금액(원) 추출."""
    if m := RE_VALUE_OK.search(text):
        return float(m.group(1)) * 1e8
    if m := RE_VALUE_MAN.search(text):
        return float(m.group(1).replace(",", "")) * 1e4
    if m := RE_VALUE_WON.search(text):
        return float(m.group(1).replace(",", ""))
    return None


def _parse_quantity(text: str) -> float | None:
    if m := RE_QUANTITY.search(text):
        return float(m.group(1).replace(",", ""))
    return None


def _parse_return(text: str) -> float | None:
    # 수익률 = 마지막 괄호 안 % 우선. 채권은 손익금이라 % 아님.
    matches = RE_RETURN.findall(text)
    return float(matches[-1]) if matches else None


def _load_ticker_map() -> dict[str, tuple[str, str]]:
    """ticker_map.csv → {정규화된 이름: (ticker, asset_type)}"""
    if not TICKER_MAP_PATH.exists():
        return {}
    df = pd.read_csv(TICKER_MAP_PATH)
    out: dict[str, tuple[str, str]] = {}
    for _, r in df.iterrows():
        key = _norm(str(r["name"]))
        out[key] = (str(r["ticker"]), str(r.get("asset_type", "stock")))
    return out


def _norm(s: str) -> str:
    """이름 정규화 — 공백 제거, 소문자."""
    return re.sub(r"\s+", "", s).lower().strip()


def _match_ticker(name: str, tmap: dict[str, tuple[str, str]]) -> tuple[str | None, str]:
    """이름 → (ticker, asset_type). 매칭 실패 시 (None, 'unknown').

    부분일치는 양방향 모두 일정 길이(8자) 이상일 때만 허용 — 짧은 키가
    긴 이름에 잘못 들러붙는 것 방지.
    """
    key = _norm(name)
    if key in tmap:
        return tmap[key]
    # 부분일치: 후보 중 가장 긴 매칭 선택, 한쪽이 8자 이상
    best = None
    best_len = 0
    for k, v in tmap.items():
        if len(k) < 8:
            continue
        if k in key or key in k:
            overlap = min(len(k), len(key))
            if overlap > best_len:
                best = v
                best_len = overlap
    return best if best else (None, "unknown")


CASH_KEYWORDS = ("이율보증형", "현금성자산", "RP", "예금", "MMF", "예치금")
FUND_KEYWORDS = ("TDF", "투자신탁", "재간접", "투자형", "혼합형")


def _classify_unmatched(name: str) -> str:
    if any(k in name for k in CASH_KEYWORDS):
        return "cash"
    if any(k in name for k in FUND_KEYWORDS):
        return "fund"
    return "unknown"


def _parse_line_tradeable(line: str, no: int, category: str, provider: str, tmap: dict) -> Holding | None:
    """주식·ETF 한 줄 파싱. 티커 매칭 실패 시 cash/fund로 폴백."""
    line = line.strip().rstrip(",")
    if not line:
        return None
    if line.startswith(("총적립액", "총평가", "합계", "총액")):
        return None

    m = re.search(r"\s+\d", line)
    name = line[: m.start()].strip() if m else line

    qty = _parse_quantity(line)
    value = _parse_value(line)
    ret = _parse_return(line)
    ticker, atype = _match_ticker(name, tmap)

    if ticker:
        return Holding(
            no=no, category=category, provider=provider, name=name, raw=line,
            quantity=qty, value_krw=value, return_pct=ret,
            ticker=ticker, asset_type=atype,
        )

    # 티커 매칭 실패 — 키워드로 자산 유형 추정
    fallback_type = _classify_unmatched(name)
    if fallback_type == "unknown":
        fallback_type = "stock"  # 진짜 모르겠으면 추적 후보로 남김(unmatched.csv로도 기록)

    return Holding(
        no=no, category=category, provider=provider, name=name, raw=line,
        quantity=qty, value_krw=value, return_pct=ret,
        ticker=None, asset_type=fallback_type,
    )


def _parse_cell_bonds(text: str, no: int, provider: str) -> list[Holding]:
    """채권 셀 — 다양한 포맷 처리."""
    out: list[Holding] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        value = _parse_value(line)
        # 종목명: 첫 숫자 앞까지 또는 미국채/국고/브라질채 다음
        m = re.search(r"\s+\d+\.\d+\s*억", line)
        if m:
            name = line[: m.start()].strip()
        else:
            name = line
        ccy = "USD" if "미국채" in name else ("BRL" if "브라질" in name else "KRW")
        out.append(Holding(
            no=no, category="채권", provider=provider, name=name, raw=line,
            value_krw=value, asset_type="bond", currency=ccy,
        ))
    return out


def _parse_cell_funds(text: str, no: int, category: str, provider: str) -> list[Holding]:
    """변액·연금 펀드 — 각 줄을 펀드 1개로. '총적립액' 등 합계 줄은 스킵."""
    out: list[Holding] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # 합계 행 식별 — 자산 표시에서 중복 방지
        if line.startswith(("총적립액", "총평가", "합계", "총액")):
            continue
        qty = _parse_quantity(line)
        value = _parse_value(line)
        ret = _parse_return(line)
        m = re.search(r"\s+\d", line)
        name = line[: m.start()].strip() if m else line
        out.append(Holding(
            no=no, category=category, provider=provider, name=name, raw=line,
            quantity=qty, value_krw=value, return_pct=ret,
            asset_type="fund",
        ))
    return out


def _parse_cell_cash(text: str, no: int, category: str, provider: str) -> list[Holding]:
    value = _parse_value(text)
    return [Holding(
        no=no, category=category, provider=provider, name=category, raw=text,
        value_krw=value, asset_type="cash",
    )]


def _parse_cell_pension(text: str, no: int, category: str, provider: str) -> list[Holding]:
    value = _parse_value(text)
    return [Holding(
        no=no, category=category, provider=provider, name=category, raw=text,
        value_krw=value, asset_type="pension",
    )]


def load_portfolio_from_excel(excel_path: Path | None = None) -> tuple[list[Holding], list[Holding]]:
    """Excel을 파싱해 (tradeable, others)로 분리.

    tradeable: yfinance로 추적할 종목 (티커 매칭된 것만)
    others: 펀드·연금·현금·매칭 실패 — 별도 표시
    """
    excel_path = excel_path or EXCEL_PATH
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 파일 없음: {excel_path}")

    tmap = _load_ticker_map()
    log.info(f"티커 매핑 로드: {len(tmap)}건")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    all_holdings: list[Holding] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            no = int(row[0])
        except (TypeError, ValueError):
            continue
        category = (row[1] or "").strip()
        provider = (row[2] or "").strip()
        body = (row[3] or "").strip()
        if not body:
            continue

        if category in TRADEABLE_CATEGORIES:
            for line in body.splitlines():
                h = _parse_line_tradeable(line, no, category, provider, tmap)
                if h:
                    all_holdings.append(h)
        elif category in BOND_CATEGORIES:
            all_holdings.extend(_parse_cell_bonds(body, no, provider))
        elif category in FUND_CATEGORIES:
            all_holdings.extend(_parse_cell_funds(body, no, category, provider))
        elif category in CASH_CATEGORIES:
            all_holdings.extend(_parse_cell_cash(body, no, category, provider))
        elif category in PENSION_CATEGORIES:
            all_holdings.extend(_parse_cell_pension(body, no, category, provider))
        else:
            # 알 수 없는 카테고리 — 펀드로 처리
            all_holdings.extend(_parse_cell_funds(body, no, category, provider))

    tradeable = [h for h in all_holdings if h.ticker]
    others = [h for h in all_holdings if not h.ticker]

    # 매칭 실패 항목 로그
    unmatched_tradeable = [h for h in all_holdings if h.asset_type in ("stock", "etf") and not h.ticker]
    if unmatched_tradeable:
        UNMATCHED_LOG.parent.mkdir(exist_ok=True)
        df = pd.DataFrame([
            {"category": h.category, "name": h.name, "raw": h.raw}
            for h in unmatched_tradeable
        ])
        df.to_csv(UNMATCHED_LOG, index=False, encoding="utf-8-sig")
        log.warning(f"티커 매칭 실패 {len(unmatched_tradeable)}건 → {UNMATCHED_LOG}")

    log.info(f"파싱 완료: 추적 {len(tradeable)}건, 비추적 {len(others)}건")
    return tradeable, others
